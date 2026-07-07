import io

from docx import Document
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from app.services.extract import classify_kind, extract_text


def make_pdf_bytes(text: str = "Hello from a test PDF") -> bytes:
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=letter)
    c.drawString(100, 700, text)
    c.save()
    return buf.getvalue()


def make_docx_bytes(text: str = "Hello from a test Word doc") -> bytes:
    doc = Document()
    doc.add_paragraph(text)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def make_xlsx_bytes(value: str = "Hello from a test spreadsheet") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = value
    ws["B1"] = 42
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_classify_kind_by_extension_and_mimetype():
    assert classify_kind("photo.png", None) == "image"
    assert classify_kind("scan.pdf", None) == "pdf"
    assert classify_kind("report.docx", None) == "docx"
    assert classify_kind("data.xlsx", None) == "xlsx"
    assert classify_kind("weird", "image/jpeg") == "image"
    assert classify_kind("mystery.bin", None) == "other"


def test_extract_pdf_text():
    text = extract_text(make_pdf_bytes("Unique PDF marker XYZ123"), "pdf")
    assert text is not None
    assert "Unique PDF marker XYZ123" in text


def test_extract_docx_text():
    text = extract_text(make_docx_bytes("Unique Docx marker ABC789"), "docx")
    assert text is not None
    assert "Unique Docx marker ABC789" in text


def test_extract_xlsx_text():
    text = extract_text(make_xlsx_bytes("Unique Xlsx marker QWE456"), "xlsx")
    assert text is not None
    assert "Unique Xlsx marker QWE456" in text
    assert "42" in text


def test_extract_text_returns_none_for_image_and_other():
    assert extract_text(b"not really an image", "image") is None
    assert extract_text(b"whatever", "other") is None


def test_extract_text_malformed_file_does_not_raise():
    assert extract_text(b"this is not a valid pdf", "pdf") is None
    assert extract_text(b"this is not a valid docx", "docx") is None
    assert extract_text(b"this is not a valid xlsx", "xlsx") is None
